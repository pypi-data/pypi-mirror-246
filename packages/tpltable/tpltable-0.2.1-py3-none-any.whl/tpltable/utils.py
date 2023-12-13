from tpltable.box4 import Box4, Vec4, Vec4Zero

##
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.cell_style import StyleArray
from openpyxl.cell.cell import MergedCell
from typing import Union, Callable, Sized, Iterable

#
import re
import os
import copy
import logging
import inspect


class ClassicalDecorator:
    def __init__(self, func):
        self.func = func

    def __call__(self, *args, **kwargs):
        return self.func(*args, **kwargs)


# overwrite load_workbook
def load_workbook(filepath, read_only=False, keep_vba=False, data_only=False, keep_links=True):
    """
    重写load_workbook，为读取到的Workbook对象添加一个name属性，用于标记该Workbook的文件名
    :param filepath:
    :param read_only:
    :param keep_vba:
    :param data_only:
    :param keep_links:
    :return:
    """
    # check filepath
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File {filepath} not found.")
    if not os.path.isfile(filepath):
        raise TypeError(f"{filepath} is not a file.")
    # END check filepath
    full_path = os.path.abspath(filepath)
    fname = os.path.basename(full_path)
    fname, _ = os.path.splitext(fname)

    book = openpyxl.load_workbook(filepath, read_only, keep_vba, data_only, keep_links)
    book.name = fname

    return book


def get_book_name(book: Workbook):
    """
    获取book的文件名
    :param book:
    :return:
    """
    _ = getattr(book, 'name', f"Workbook{id(book)}")
    return _


# use info, warn, error
# output like:
# print to console(but do not contain time)
#   [WHO] - INFO - info message
#   [WHO] - WARN - warn message
#   [WHO] - !ERRORType - error message
##
class TplLogger:
    def __init__(self, name='tpltable', level=logging.INFO):
        self.logger = logging.getLogger(name)
        self.logger.setLevel(level)
        self.formatter = logging.Formatter('%(message)s')
        self.console_handler = logging.StreamHandler()
        self.console_handler.setLevel(level)
        self.console_handler.setFormatter(self.formatter)
        self.logger.addHandler(self.console_handler)

    def set_level(self, level):
        self.logger.setLevel(level)
        self.console_handler.setLevel(level)

    def info(self, *msg):
        msg = ' '.join([str(_) for _ in msg])
        _caller = self._get_caller()
        self.logger.info('[{}]'.format(_caller).ljust(25) + '[INFO]'.ljust(20) + msg)

    def warn(self, *msg):
        msg = ' '.join([str(_) for _ in msg])
        _caller = self._get_caller()
        self.logger.warning('[{}]'.format(_caller).ljust(25) + '[WARN]'.ljust(20) + msg)

    def error(self, err: type, *msg, exit=False):
        msg = ' '.join([str(_) for _ in msg])
        _caller = self._get_caller()
        self.logger.error('[{}]'.format(_caller).ljust(25) + '[!{}]'.format(err.__name__).ljust(20) + msg)
        if exit:
            raise err(msg)

    def _get_caller(self):
        _stacks = inspect.stack()
        _frame = _stacks[2]
        _module = inspect.getmodule(_frame[0])
        _caller = f"{_module.__name__}.{_frame[3]}"
        return _caller


log = TplLogger()


def sheet_sizeof(sheet):
    """
    获取sheet的大小
    并不能单纯的使用sheet.max_row和sheet.max_column，因为这两个属性只能获取到有值的最大行和列（甚至略有修改属性都计算在内）
    我们需要获取的是sheet的cell有内容的最大行和列
    *从sheet.max_row和sheet.max_column的像内逼近，直到找到第一个有值的cell
    :param sheet:
    :return:
    """
    return sheet.max_row, sheet.max_column
    if hasattr(sheet, '_max_row_') and hasattr(sheet, '_max_col_'):
        return sheet._max_row_, sheet._max_col_

    row_flag, col_flag = False, False
    max_row, max_col = sheet.max_row, sheet.max_column

    # row:
    for row in range(max_row, 0, -1):
        for col in range(1, max_col + 1):
            if sheet.cell(row=row, column=col).value is not None:
                max_row, row_flag = row, True
                break
        if row_flag:
            break

    # col
    for col in range(max_col, 0, -1):
        for row in range(1, max_row + 1):
            if sheet.cell(row=row, column=col).value is not None:
                max_col, col_flag = col, True
                break
        if col_flag:
            break

    # return max_row, max_col
    setattr(sheet, '_max_row_', max_row)
    setattr(sheet, '_max_col_', max_col)

    return max_row, max_col



_styles_attrs = {
    '_fonts': 'fontId',
    '_fills': 'fillId',
    '_borders': 'borderId',
    '_protections': 'protectionId',
    '_alignments': 'alignmentId',
    '_number_formats': 'numFmtId',
    '_named_styles': 'xfId',
    # pivotButton?
    # quotePrefix?
}


def share_style(book_source: Workbook, book_target: Workbook):
    """
    将book_source的样式，复制到book_target.
    * 首次调用时，会在book_target中创建一个_SHARE_WITH属性，用于标记book_target已经和book_source share了style
    :param book_source:
    :param book_target:
    :return:
    """
    if getattr(book_target, '_SHARE_WITH', None) == id(book_source):
        return
    for style in _styles_attrs:
        coll_source = getattr(book_source, style)
        setattr(book_target, style, coll_source)
    book_target._SHARE_WITH = id(book_source)


def copy_style(src_cell, dest_cell):
    """
    复制源单元格的样式到目标单元格。
    :param src_cell: 源单元格
    :param dest_cell: 目标单元格
    """
    # if src_cell.has_style:
    #     dest_cell.font = copy.copy(src_cell.font)
    if src_cell.has_style:
        share_style(src_cell.parent.parent, dest_cell.parent.parent)
        if not getattr(dest_cell, "_style"):
            dest_cell._style = StyleArray()
        for style, attr in _styles_attrs.items():
            _value = getattr(src_cell._style, attr, None)
            if _value is not None:
                setattr(dest_cell._style, attr, _value)


def check_ndarr_dict(ndarr_dict: np.ndarray, check_element=True, exit=True) -> bool:
    """
    检查ndarr_dict的合法性. 如果不合法，会抛出异常
    :param ndarr_dict:
    :param check_element: 是否检查ndarr_dict中每个元素的合法性
    :param exit: 是否抛出异常
    :return: True or False
    """
    if not isinstance(ndarr_dict, np.ndarray):
        log.error(ValueError, 'ndarr_dict should be a np.ndarray, but got {}'.format(type(ndarr_dict)), exit=exit)
        return False
    if ndarr_dict.dtype != object:
        log.error(ValueError, 'ndarr_dict should be a np.ndarray with dtype object, but got {}'.format(ndarr_dict.dtype), exit=exit)
        return False
    if len(ndarr_dict.shape) != 2:
        log.error(ValueError, 'ndarr_dict should be a np.ndarray with shape(ah, aw), but got shape{}'.format(ndarr_dict.shape), exit=exit)
        return False

    # check dict
    if check_element:
        for ih in range(ndarr_dict.shape[0]):
            for iw in range(ndarr_dict.shape[1]):
                _ = ndarr_dict[ih, iw]
                if not isinstance(_, dict):
                    log.error(ValueError, 'ndarr_dict[{}, {}] should be a dict, but got {}'.format(ih, iw, type(_)), exit=exit)
                    return False
                for key, s in _.items():
                    if not isinstance(s, str):
                        log.error(ValueError, 'ndarr_dict[{}, {}][{}] should be a str, but got {}'.format(ih, iw, key, type(s)), exit=exit)
                    return False
    else:
        # 检查一个元素即可
        if ndarr_dict.shape[0] == 0 or ndarr_dict.shape[1] == 0:
            return True
        _ = ndarr_dict[0, 0]
        if not isinstance(_, dict):
            log.error(ValueError, 'ndarr_dict[0, 0] should be a dict, but got {}'.format(type(_)), exit=exit)
            return False
        for key, s in _.items():
            if not isinstance(s, str):
                log.error(ValueError, 'ndarr_dict[0, 0][{}] should be a str, but got {}'.format(key, type(s)), exit=exit)
                return False
            break
    return True


def copy_area(src_sheet: Worksheet, dest_sheet: Worksheet, src_area: Vec4, dest_area: Vec4):
    """
    将src_sheet中的src_area复制到dest_sheet中的dest_area
    :param src_sheet:
    :param dest_sheet:
    :param src_area:
    :param dest_area:
    :return:
    """
    assert src_area.h == dest_area.h and src_area.w == dest_area.w
    for row in range(src_area.h):
        for col in range(src_area.w):
            src_cell = src_sheet.cell(row=src_area.row + row, column=src_area.col + col)
            dest_cell = dest_sheet.cell(row=dest_area.row + row, column=dest_area.col + col)
            # if isinstance(src_cell, MergedCell):
            #     src_cell = src_sheet[src_cell.coordinate]
            # if isinstance(dest_cell, MergedCell):
            #     dest_cell = dest_sheet[dest_cell.coordinate]
            dest_cell.value = src_cell.value
            copy_style(src_cell, dest_cell)


def tpl_keysplit(s, advance=True):
    """
    严格的按照$分割字符串,要求$后必须有字符.
    支持以$XX:YY的形式
    例如:
        $a$b -> ["$a", "$b"]
        $a:b$c -> ["$a:b", "$c"]
        $a$$b -> Raise ValueError
        $a -> ["$a"]
        '' -> []
    :param s:
    :return:
    """
    _ = re.findall(r'\$[^\$]+', s)
    # 检查能否还原
    _s = ''.join(_)
    if _s != s:
        # raise ValueError(f"Unexpected string: {s} (Maybe you mean: {_s} ?)")
        log.error(ValueError, f"Unexpected string: {s} (Maybe you mean: {_s} ?)", exit=True)
    # 检查每个key是否为$XX或$XX:YY的形式，其中XX和YY都是类似python变量名的形式
    for key in _:
        if not advance:
            rc = re.match(r'^\$[a-zA-Z_][a-zA-Z0-9_]*$', key)
        else:
            rc = re.match(r'^\$[a-zA-Z_][a-zA-Z0-9_&|]*(\:[a-zA-Z_][a-zA-Z0-9_]*)?$', key)
        if not rc:
            log.error(ValueError, f"Unexpected key: {key}", exit=True)

    # END Check

    return _


def bound_keys(tpl, area: Vec4):
    """
    根据vec4指定的范围，将tpl中位于区域内的key找出来
    :param tpl:
    :param area: Vec4
    :return: dict
    """
    _ = {}
    for key, v4 in tpl.items():
        if area.contain(v4):
            _[key] = v4
    return _


def expand_area_key(key, tpl: dict):
    """
    将area_key扩展为一个列表
    :param key:
    :param tpl: $XXX:vec4
    :return:
    """
    ## find : and split
    _find = key.find(':')
    if _find == -1:
        return [key]

    ## split
    _a, _b = key[:_find], key[_find + 1:]
    if _b.startswith('$'):
        log.warn(f"Unexpected key: {key}. Auto ignore.")
    else:
        _b = '$' + _b

    ## get vec4
    _v4_a, _v4_b = tpl.get(_a, None), tpl.get(_b, None)
    if _v4_a is None:
        log.error(ValueError, f"Unexpected key: {_a}. Failed find in tpl-format: {tpl}", exit=True)
    if _v4_b is None:
        log.error(ValueError, f"Unexpected key: {_b}. Failed find in tpl-format: {tpl}", exit=True)
    if not isinstance(_v4_a, Vec4):
        log.error(ValueError, f"Unexpected key: {_a}:{_v4_a}. Unexpected value found in tpl-format: {tpl}", exit=True)
    if not isinstance(_v4_b, Vec4):
        log.error(ValueError, f"Unexpected key: {_b}:{_v4_b}. Unexpected value found in tpl-format: {tpl}", exit=True)

    # Check Zero
    if _v4_a.iszero():
        log.error(ValueError, f"Unexpected key: {_a}:{_v4_a}. {_a} maybe a key defined in pipe.", exit=True)
    if _v4_b.iszero():
        log.error(ValueError, f"Unexpected key: {_b}:{_v4_b}. {_b} maybe a key defined in pipe.", exit=True)

    ## bounds
    # create a new v4: from _a to _b
    min_x, min_y = min(_v4_a.x, _v4_b.x), min(_v4_a.y, _v4_b.y)
    max_x, max_y = (_v4_a.x + _v4_a.w) if _v4_a.x > _v4_b.x else (_v4_b.x + _v4_b.w), \
        (_v4_a.y + _v4_a.h) if _v4_a.y > _v4_b.y else (_v4_b.y + _v4_b.h)
    _v4 = Vec4(min_y, min_x, max_y - min_y, max_x - min_x)
    # bound
    _tar_tpl = bound_keys(tpl, _v4)

    return list(_tar_tpl.keys())


if __name__ == '__main__':
    print(
        tpl_keysplit('$a$b'),
        tpl_keysplit('$a:b$c'),
        tpl_keysplit('$a'),
        tpl_keysplit(''),
    )

    # print(tpl_keysplit('$a$$b'))
