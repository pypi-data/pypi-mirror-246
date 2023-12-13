from openpyxl.utils import get_column_letter, range_boundaries
class _Vec4_basic:
    """
    A class to represent a 4D vector.
    Specifically, this class is used to represent a bounding box:
        row(y), # r y
        col(x), # c x
        height(h), # h
        width(w) # w
    """
    def __init__(self, y, x, h, w):
        self.y = y
        self.x = x
        self.h = h
        self.w = w

    @property
    def r(self):
        return self.y + 1

    @property
    def c(self):
        return self.x + 1

    @property
    def row(self):
        return self.y + 1

    @property
    def col(self):
        return self.x + 1

    @property
    def height(self):
        return self.h

    @property
    def width(self):
        return self.w

    def move(self, x, y):
        return Vec4(self.y + y, self.x + x, self.h, self.w)

    def overlap(self, other):
        """
        判断两个矩形是否重叠
        """
        return self.x < other.x + other.w and self.x + self.w > other.x and self.y < other.y + other.h and self.y + self.h > other.y

    def contain(self, other):
        """
        判断一个矩形是否包含另一个矩形
        """
        return self.x <= other.x and self.x + self.w >= other.x + other.w and self.y <= other.y and self.y + self.h >= other.y + other.h

    def iszero(self):
        return self.x == 0 and self.y == 0 and self.w == 0 and self.h == 0

class _Vec4_magic(_Vec4_basic):
    """
    A class to represent a 4D vector.
    Specifically, this class is used to represent a bounding box:
        row(y), # r y
        col(x), # c x
        height(h), # h
        width(w) # w
    """
    def __repr__(self):
        return f"Vec4(y={self.y}, x={self.x}, h={self.h}, w={self.w})"

    def __str__(self):
        return f"Vec4(y={self.y}, x={self.x}, h={self.h}, w={self.w})"

    def __eq__(self, other):
        return self.y == other.y and self.x == other.x and self.h == other.h and self.w == other.w

    def __ne__(self, other):
        return not self.__eq__(other)

    def __add__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y + other.y, self.x + other.x, self.h + other.h, self.w + other.w)
        else:
            return Vec4(self.y + other, self.x + other, self.h + other, self.w + other)

    def __sub__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y - other.y, self.x - other.x, self.h - other.h, self.w - other.w)
        else:
            return Vec4(self.y - other, self.x - other, self.h - other, self.w - other)

    def __mul__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y * other.y, self.x * other.x, self.h * other.h, self.w * other.w)
        else:
            return Vec4(self.y * other, self.x * other, self.h * other, self.w * other)

    def __truediv__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y / other.y, self.x / other.x, self.h / other.h, self.w / other.w)
        else:
            return Vec4(self.y / other, self.x / other, self.h / other, self.w / other)

    def __floordiv__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y // other.y, self.x // other.x, self.h // other.h, self.w // other.w)
        else:
            return Vec4(self.y // other, self.x // other, self.h // other, self.w // other)

    def __mod__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y % other.y, self.x % other.x, self.h % other.h, self.w % other.w)
        else:
            return Vec4(self.y % other, self.x % other, self.h % other, self.w % other)

    def __pow__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y ** other.y, self.x ** other.x, self.h ** other.h, self.w ** other.w)
        else:
            return Vec4(self.y ** other, self.x ** other, self.h ** other, self.w ** other)

    def __and__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y & other.y, self.x & other.x, self.h & other.h, self.w & other.w)
        else:
            return Vec4(self.y & other, self.x & other, self.h & other, self.w & other)

    def __or__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y | other.y, self.x | other.x, self.h | other.h, self.w | other.w)
        else:
            return Vec4(self.y | other, self.x | other, self.h | other, self.w | other)

    def __xor__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y ^ other.y, self.x ^ other.x, self.h ^ other.h, self.w ^ other.w)
        else:
            return Vec4(self.y ^ other, self.x ^ other, self.h ^ other, self.w ^ other)

    def __lshift__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y << other.y, self.x << other.x, self.h << other.h, self.w << other.w)
        else:
            return Vec4(self.y << other, self.x << other, self.h << other, self.w << other)

    def __rshift__(self, other):
        if isinstance(other, Vec4):
            return Vec4(self.y >> other.y, self.x >> other.x, self.h >> other.h, self.w >> other.w)
        else:
            return Vec4(self.y >> other, self.x >> other, self.h >> other, self.w >> other)

    def __abs__(self):
        return Vec4(abs(self.y), abs(self.x), abs(self.h), abs(self.w))

    def __neg__(self):
        return Vec4(-self.y, -self.x, -self.h, -self.w)

    def __pos__(self):
        return Vec4(+self.y, +self.x, +self.h, +self.w)

    def __invert__(self):
        return Vec4(~self.y, ~self.x, ~self.h, ~self.w)

    def __lt__(self, other):
        return self.y < other.y and self.x < other.x and self.h < other.h and self.w < other.w

    def __le__(self, other):
        return self.y <= other.y and self.x <= other.x and self.h <= other.h and self.w <= other.w

    def __gt__(self, other):
        return self.y > other.y and self.x > other.x and self.h > other.h and self.w > other.w

    def __ge__(self, other):
        return self.y >= other.y and self.x >= other.x and self.h >= other.h and self.w >= other.w

    def __len__(self):
        return 4

    def __getitem__(self, item):
        if isinstance(item, int):
            if item == 0:
                return self.y
            elif item == 1:
                return self.x
            elif item == 2:
                return self.h
            elif item == 3:
                return self.w
            else:
                raise IndexError("Vec4 index out of range")
        elif isinstance(item, slice):
            if item.start is None:
                start = 0
            else:
                start = item.start
            if item.stop is None:
                stop = 4
            else:
                stop = item.stop
            if item.step is None:
                step = 1
            else:
                step = item.step
            return [self[i] for i in range(start, stop, step)]
        else:
            raise TypeError("Vec4 indices must be integers or slices, not " + str(type(item)))

    def __setitem__(self, key, value):
        if isinstance(key, int):
            if key == 0:
                self.y = value
            elif key == 1:
                self.x = value
            elif key == 2:
                self.h = value
            elif key == 3:
                self.w = value
            else:
                raise IndexError("Vec4 index out of range")
        elif isinstance(key, slice):
            if key.start is None:
                start = 0
            else:
                start = key.start
            if key.stop is None:
                stop = 4
            else:
                stop = key.stop
            if key.step is None:
                step = 1
            else:
                step = key.step
            for i in range(start, stop, step):
                self[i] = value
        else:
            raise TypeError("Vec4 indices must be integers or slices, not " + str(type(key)))


    def __iadd__(self, other):
        if isinstance(other, Vec4):
            self.y += other.y
            self.x += other.x
            self.h += other.h
            self.w += other.w
        else:
            self.y += other
            self.x += other
            self.h += other
            self.w += other
        return self

    def __isub__(self, other):
        if isinstance(other, Vec4):
            self.y -= other.y
            self.x -= other.x
            self.h -= other.h
            self.w -= other.w
        else:
            self.y -= other
            self.x -= other
            self.h -= other
            self.w -= other
        return self

    def __imul__(self, other):
        if isinstance(other, Vec4):
            self.y *= other.y
            self.x *= other.x
            self.h *= other.h
            self.w *= other.w
        else:
            self.y *= other
            self.x *= other
            self.h *= other
            self.w *= other
        return self

    def __itruediv__(self, other):
        if isinstance(other, Vec4):
            self.y /= other.y
            self.x /= other.x
            self.h /= other.h
            self.w /= other.w
        else:
            self.y /= other
            self.x /= other
            self.h /= other
            self.w /= other
        return self

    def __ifloordiv__(self, other):
        if isinstance(other, Vec4):
            self.y //= other.y
            self.x //= other.x
            self.h //= other.h
            self.w //= other.w
        else:
            self.y //= other
            self.x //= other
            self.h //= other
            self.w //= other
        return self

    def __imod__(self, other):
        if isinstance(other, Vec4):
            self.y %= other.y
            self.x %= other.x
            self.h %= other.h
            self.w %= other.w
        else:
            self.y %= other
            self.x %= other
            self.h %= other
            self.w %= other
        return self

    def __ipow__(self, other):
        if isinstance(other, Vec4):
            self.y **= other.y
            self.x **= other.x
            self.h **= other.h
            self.w **= other.w
        else:
            self.y **= other
            self.x **= other
            self.h **= other
            self.w **= other
        return self

    def __iand__(self, other):
        if isinstance(other, Vec4):
            self.y &= other.y
            self.x &= other.x
            self.h &= other.h
            self.w &= other.w
        else:
            self.y &= other
            self.x &= other
            self.h &= other
            self.w &= other
        return self

    def __ior__(self, other):
        if isinstance(other, Vec4):
            self.y |= other.y
            self.x |= other.x
            self.h |= other.h
            self.w |= other.w
        else:
            self.y |= other
            self.x |= other
            self.h |= other
            self.w |= other
        return self

    def __ixor__(self, other):
        if isinstance(other, Vec4):
            self.y ^= other.y
            self.x ^= other.x
            self.h ^= other.h
            self.w ^= other.w
        else:
            self.y ^= other
            self.x ^= other
            self.h ^= other
            self.w ^= other
        return self

    def __ilshift__(self, other):
        if isinstance(other, Vec4):
            self.y <<= other.y
            self.x <<= other.x
            self.h <<= other.h
            self.w <<= other.w
        else:
            self.y <<= other
            self.x <<= other
            self.h <<= other
            self.w <<= other
        return self

    def __irshift__(self, other):
        if isinstance(other, Vec4):
            self.y >>= other.y
            self.x >>= other.x
            self.h >>= other.h
            self.w >>= other.w
        else:
            self.y >>= other
            self.x >>= other
            self.h >>= other
            self.w >>= other
        return self

    def __bool__(self):
        return bool(self.y or self.x or self.h or self.w)

    def __contains__(self, item):
        return item in [self.y, self.x, self.h, self.w]

    def __hash__(self):
        return hash((self.y, self.x, self.h, self.w))

    def __round__(self, n=None):
        return Vec4(round(self.y, n), round(self.x, n), round(self.h, n), round(self.w, n))

    def __floor__(self):
        return Vec4(math.floor(self.y), math.floor(self.x), math.floor(self.h), math.floor(self.w))

    def __ceil__(self):
        return Vec4(math.ceil(self.y), math.ceil(self.x), math.ceil(self.h), math.ceil(self.w))

    def __trunc__(self):
        return Vec4(math.trunc(self.y), math.trunc(self.x), math.trunc(self.h), math.trunc(self.w))

    def __copy__(self):
        return Vec4(self.y, self.x, self.h, self.w)

    def __deepcopy__(self, memodict={}):
        return Vec4(self.y, self.x, self.h, self.w)

    def __iter__(self):
        return iter([self.y, self.x, self.h, self.w])

    def __reversed__(self):
        return reversed([self.y, self.x, self.h, self.w])

    def __index__(self):
        return self.y + self.x + self.h + self.w

    def __oct__(self):
        return oct(self.y + self.x + self.h + self.w)

    def __hex__(self):
        return hex(self.y + self.x + self.h + self.w)

    def __int__(self):
        return int(self.y + self.x + self.h + self.w)

    def __float__(self):
        return float(self.y + self.x + self.h + self.w)

    # 允许按照 .xy .xywh .rc这样的方式访问
    @property
    def xy(self):
        return self.x, self.y

    @property
    def rc(self):
        return self.r, self.c

    @property
    def wh(self):
        return self.w, self.h

    @property
    def xywh(self):
        return self.x, self.y, self.w, self.h

    @property
    def rcwh(self):
        return self.r, self.c, self.w, self.h


class Vec4(_Vec4_magic):
    """
    A class to represent a 4D vector.
    Specifically, this class is used to represent a bounding box:
        row(y), # r y
        col(x), # c x
        height(h), # h
        width(w) # w
    """
    @property
    def range4(self) -> list:
        """
        Return a openpyxl.range4 from the 4D vector.
        (min_col, min_row, max_col, max_row)
        * range4 starts from 1
        * range4 includes both min and max, which means [min, max].
        """
        return [self.col, self.row, self.x + self.w, self.y + self.h]

    @staticmethod
    def from_range4(range4):
        """
        Return a 4D vector from a openpyxl.range4.
        (min_col, min_row, max_col, max_row)
        * range4 starts from 1
        * range4 includes both min and max, which means [min, max].
        """
        return Vec4(range4[1] - 1, range4[0] - 1, range4[3] - range4[1] + 1, range4[2] - range4[0] + 1)

    @property
    def letter(self) -> str:
        """
        Return a letter from the 4D vector.
        get the excel letter from the 4D vector.
        """
        range4 = self.range4
        # openpyxl.utils
        # get letter like "A?:X?"
        return get_column_letter(range4[0]) + str(range4[1]) + ":" + get_column_letter(range4[2]) + str(range4[3])

    @staticmethod
    def from_letter(letter):
        """
        Return a 4D vector from a letter.
        get the 4D vector from the excel letter.
        """
        # use openpyxl.utils
        # get letter like "X?" and "A?:X?"
        # from letters to range4
        range4 = range_boundaries(letter)
        return Vec4.from_range4(range4)


def Vec4Zero():
    return Vec4(0, 0, 0, 0)

class Box4(Vec4):
    """
    一个可以装数据的盒子，可以装入任何数据。
    同时继承Vec4，表示了数据的位置。
    *value 不参与四则运算
    """
    def __init__(self, y, x, h, w, value=None):
        super().__init__(y, x, h, w)
        self.value = value

    def __str__(self):
        return "Box4: x={}, y={}, w={}, h={}, data={}".format(self.x, self.y, self.w, self.h, self.value)

    def __repr__(self):
        return "Box4: x={}, y={}, w={}, h={}, data={}".format(self.x, self.y, self.w, self.h, self.value)

    def __eq__(self, other):
        return self.x == other.x and self.y == other.y and self.w == other.w and self.h == other.h and self.value == other.value

    def __ne__(self, other):
        return not self.__eq__(other)

    def __hash__(self):
        return hash((self.x, self.y, self.w, self.h, self.value))

    def __copy__(self):
        return Box4(self.y, self.x, self.h, self.w, self.value)

    def __deepcopy__(self, memodict={}):
        return Box4(self.y, self.x, self.h, self.w, self.value)

if __name__ == '__main__':
    v4 = Vec4.from_letter("A1:AAA999")
    print(
        v4,
        v4.range4,
        v4.letter,
        v4.xy,
        v4.wh,
    )